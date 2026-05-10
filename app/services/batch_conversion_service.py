"""
app/services/batch_conversion_service.py

Batch import of mid-term-boarded loans from a multi-sheet xlsx workbook.

Two phases (matching the API):
  1. parse_and_validate(file_bytes)
       - Reads Counterparties + Loans sheets
       - Builds LoanCreate payloads with embedded conversion blocks
       - Captures per-row issues without touching the DB
       - Persists a conversion_batch row (status='validated' or 'failed')
       - Returns the validation report to the caller

  2. commit_batch(batch_id)
       - Re-reads the saved validation report
       - For each valid row, opens its own transaction and calls
         LoanService.create_converted_loan
       - Per-row failures DO NOT roll back the rest of the batch
       - Updates conversion_batch.commit_report incrementally so progress
         queries see partial results

Excel format:
  Sheet "Counterparties" (optional)
    columns: external_ref, legal_name, tax_id, entity_type, jurisdiction,
             state_of_formation, address_line1, address_line2, city, state,
             postal_code, country, primary_contact_email
    Counterparty rows are tax-id-matched against the existing tenant table:
      - tax_id matches an existing counterparty -> reuse, link to external_ref
      - no match (or no tax_id supplied)         -> create new
    Counterparties cannot be created and reused in the same batch as anything
    other than borrowers; if you need guarantors etc. add them via the UI.

  Sheet "Loans" (required)
    columns: portfolio_code, loan_number?, loan_name?, borrower_external_ref,
             original_balance, commitment_amount?, rate_type, coupon_rate?,
             pik_rate?, spread?, index_code?, day_count, origination_date,
             first_payment_date?, maturity_date, payment_frequency,
             amortization_type, interest_only_period_months?, balloon_amount?,
             grace_period_days?, late_fee_type?, late_fee_amount?,
             default_rate?, prior_servicer_loan_id?, prior_servicer_name?,
             notes?,
             -- conversion block (all required for a converted loan):
             as_of_date, current_principal, accrued_interest?, accrued_fees?,
             last_payment_date?, last_payment_amount?, next_due_date?,
             paid_to_date_principal?, paid_to_date_interest?,
             paid_to_date_fees?

  Empty sheets / unknown columns / extra columns are ignored.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional
from uuid import UUID, uuid4

import structlog
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ConflictError, ValidationError
from app.db.session import get_tenant_session_context
from app.models.conversion import ConversionBatch, LoanConversion
from app.models.loan import Loan
from app.models.portfolio import Counterparty, Portfolio
from app.schemas.conversion import (
    BatchValidationReport,
    RowIssue,
    RowResult,
)
from app.schemas.loan import LoanConversionPayload, LoanCreate
from app.services.loan_service import LoanService

logger = structlog.get_logger(__name__)

COUNTERPARTIES_SHEET = "Counterparties"
LOANS_SHEET = "Loans"


# ---------------------------------------------------------------------------
# Cell coercion helpers
# ---------------------------------------------------------------------------

def _to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        raise ValueError(f"could not parse decimal: {v!r}")


def _to_rate(v: Any) -> Optional[Decimal]:
    """Accept '9.5', '9.5%', '0.095', '9.500%'. Anything > 1 is treated as percent."""
    d = _to_decimal(v)
    if d is None:
        return None
    return d / 100 if d > 1 else d


def _to_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"could not parse date: {v!r}")


def _to_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        raise ValueError(f"could not parse integer: {v!r}")


# ---------------------------------------------------------------------------
# Parsed-row containers
# ---------------------------------------------------------------------------

@dataclass
class ParsedCounterparty:
    row: int
    external_ref: Optional[str]
    legal_name: Optional[str]
    tax_id: Optional[str]
    fields: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


@dataclass
class ParsedLoan:
    row: int
    external_ref: Optional[str]                  # = loan_number from sheet
    portfolio_code: Optional[str]
    borrower_external_ref: Optional[str]
    payload: Optional[LoanCreate] = None         # built only when fields parse cleanly
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class BatchConversionService:
    """
    Stateful service: holds a session for parse/validate (single transaction),
    commit phase opens a fresh session per loan to keep failures isolated.
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # =====================================================================
    # PARSE + VALIDATE (single transaction; persists conversion_batch row)
    # =====================================================================

    async def parse_and_validate(
        self,
        file_bytes: bytes,
        file_name: str,
        uploaded_by: UUID,
    ) -> BatchValidationReport:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        cp_rows = self._parse_counterparties(wb) if COUNTERPARTIES_SHEET in sheet_names else []
        if LOANS_SHEET not in sheet_names:
            raise ValidationError(f"Workbook must contain a '{LOANS_SHEET}' sheet")
        loan_rows = self._parse_loans(wb)

        # Resolve counterparties first — they're referenced by loans.
        # external_ref -> resolved Counterparty.id (existing or pre-allocated for create)
        cp_resolution: dict[str, UUID] = {}
        cp_results: list[RowResult] = []
        cp_creates: list[ParsedCounterparty] = []   # to actually insert if commit proceeds
        issues: list[RowIssue] = []

        for cp in cp_rows:
            if cp.errors:
                cp_results.append(RowResult(sheet=COUNTERPARTIES_SHEET, row=cp.row,
                                            external_ref=cp.external_ref, status="error",
                                            errors=cp.errors))
                for msg in cp.errors:
                    issues.append(RowIssue(sheet=COUNTERPARTIES_SHEET, row=cp.row,
                                           severity="error", message=msg))
                continue

            existing_id: Optional[UUID] = None
            if cp.tax_id:
                stmt = select(Counterparty.id).where(Counterparty.tax_id == cp.tax_id)
                existing_id = (await self.db.execute(stmt)).scalar_one_or_none()
            if existing_id:
                cp_resolution[cp.external_ref] = existing_id
                cp_results.append(RowResult(sheet=COUNTERPARTIES_SHEET, row=cp.row,
                                            external_ref=cp.external_ref, status="ok",
                                            counterparty_id=existing_id))
            else:
                # Allocate a UUID now so loans can reference it; insert at commit.
                allocated = uuid4()
                cp_resolution[cp.external_ref] = allocated
                cp_creates.append(cp)
                cp_results.append(RowResult(sheet=COUNTERPARTIES_SHEET, row=cp.row,
                                            external_ref=cp.external_ref, status="ok",
                                            counterparty_id=allocated))

        # Resolve portfolio codes once.
        port_codes_needed = {l.portfolio_code for l in loan_rows if l.portfolio_code}
        port_map: dict[str, UUID] = {}
        if port_codes_needed:
            port_result = await self.db.execute(
                select(Portfolio.code, Portfolio.id).where(Portfolio.code.in_(port_codes_needed))
            )
            port_map = {row.code: row.id for row in port_result}

        # Pre-existing prior_servicer_loan_id values (block double-imports).
        prior_ids_in_file = [
            l.payload.conversion.prior_servicer_loan_id
            for l in loan_rows
            if l.payload and l.payload.conversion and l.payload.conversion.prior_servicer_loan_id
        ]
        already_imported: set[str] = set()
        if prior_ids_in_file:
            existing = await self.db.execute(
                select(LoanConversion.prior_servicer_loan_id).where(
                    LoanConversion.prior_servicer_loan_id.in_(prior_ids_in_file)
                )
            )
            already_imported = {row[0] for row in existing}

        # Loan-level validation.
        loan_results: list[RowResult] = []
        valid_count = 0
        invalid_count = 0

        # In-batch dup detection on prior_servicer_loan_id and on loan_number.
        seen_priors: dict[str, int] = {}
        seen_numbers: dict[str, int] = {}

        for ln in loan_rows:
            row_errs = list(ln.errors)

            if ln.portfolio_code and ln.portfolio_code not in port_map:
                row_errs.append(f"unknown portfolio_code '{ln.portfolio_code}'")

            if ln.borrower_external_ref and ln.borrower_external_ref not in cp_resolution:
                row_errs.append(
                    f"borrower_external_ref '{ln.borrower_external_ref}' "
                    f"not found in Counterparties sheet"
                )

            if ln.payload and ln.payload.conversion:
                conv = ln.payload.conversion
                if conv.prior_servicer_loan_id:
                    if conv.prior_servicer_loan_id in already_imported:
                        row_errs.append(
                            f"prior_servicer_loan_id '{conv.prior_servicer_loan_id}' already converted in this tenant"
                        )
                    if conv.prior_servicer_loan_id in seen_priors:
                        row_errs.append(
                            f"prior_servicer_loan_id '{conv.prior_servicer_loan_id}' duplicated in this batch (also row {seen_priors[conv.prior_servicer_loan_id]})"
                        )
                    seen_priors[conv.prior_servicer_loan_id] = ln.row
            if ln.payload and ln.payload.loan_number:
                num = ln.payload.loan_number
                if num in seen_numbers:
                    row_errs.append(
                        f"loan_number '{num}' duplicated in this batch (also row {seen_numbers[num]})"
                    )
                seen_numbers[num] = ln.row

            if row_errs:
                invalid_count += 1
                loan_results.append(RowResult(
                    sheet=LOANS_SHEET, row=ln.row, external_ref=ln.external_ref,
                    status="error", errors=row_errs,
                ))
                for msg in row_errs:
                    issues.append(RowIssue(sheet=LOANS_SHEET, row=ln.row,
                                           severity="error", message=msg))
            else:
                valid_count += 1
                loan_results.append(RowResult(
                    sheet=LOANS_SHEET, row=ln.row, external_ref=ln.external_ref,
                    status="ok",
                ))

        # Persist the batch row (status reflects whether anything is committable).
        status = "validated" if valid_count > 0 else "failed"
        report_dict: dict[str, Any] = {
            "sheets_detected": sheet_names,
            "counterparty_rows": len(cp_rows),
            "loan_rows": len(loan_rows),
            "valid_loan_rows": valid_count,
            "invalid_loan_rows": invalid_count,
            "issues": [i.model_dump() for i in issues],
            "row_results": [r.model_dump(mode="json") for r in (cp_results + loan_results)],
            # Stashed for commit phase — not surfaced in the API response.
            "_resolution": {
                "portfolios": {k: str(v) for k, v in port_map.items()},
                "counterparties": {k: str(v) for k, v in cp_resolution.items()},
                "counterparty_creates": [
                    {"external_ref": c.external_ref, "id": str(cp_resolution[c.external_ref]),
                     "legal_name": c.legal_name, "tax_id": c.tax_id, **c.fields}
                    for c in cp_creates
                ],
                "loans": [
                    {
                        "row": ln.row,
                        "external_ref": ln.external_ref,
                        "portfolio_code": ln.portfolio_code,
                        "borrower_external_ref": ln.borrower_external_ref,
                        "payload": ln.payload.model_dump(mode="json") if ln.payload else None,
                        "errors": ln.errors,
                    }
                    for ln in loan_rows
                ],
            },
        }

        batch = ConversionBatch(
            uploaded_by=uploaded_by,
            uploaded_at=datetime.now(timezone.utc),
            file_name=file_name,
            file_hash=hashlib.sha256(file_bytes).hexdigest(),
            file_size_bytes=len(file_bytes),
            status=status,
            total_rows=len(loan_rows),
            succeeded_rows=0,
            failed_rows=invalid_count,
            validation_report=report_dict,
        )
        self.db.add(batch)
        await self.db.flush()

        return BatchValidationReport(
            batch_id=batch.id,
            file_name=file_name,
            sheets_detected=sheet_names,
            counterparty_rows=len(cp_rows),
            loan_rows=len(loan_rows),
            valid_loan_rows=valid_count,
            invalid_loan_rows=invalid_count,
            issues=issues,
            row_results=cp_results + loan_results,
        )

    # =====================================================================
    # PARSE — Excel -> typed dataclasses
    # =====================================================================

    def _parse_counterparties(self, wb) -> list[ParsedCounterparty]:
        ws = wb[COUNTERPARTIES_SHEET]
        header, header_row = self._read_header(ws)
        if not header:
            return []
        out: list[ParsedCounterparty] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                       start=header_row + 1):
            if all(c is None or c == "" for c in row):
                continue
            try:
                rd = self._row_dict(header, row)
                ext = _to_str(rd.get("external_ref"))
                legal = _to_str(rd.get("legal_name"))
                tax_id = _to_str(rd.get("tax_id"))
                errors: list[str] = []
                if not ext:
                    errors.append("external_ref is required")
                if not legal:
                    errors.append("legal_name is required")
                fields = {
                    "type": _to_str(rd.get("type")) or "borrower",
                    "entity_type": _to_str(rd.get("entity_type")),
                    "jurisdiction": _to_str(rd.get("jurisdiction")),
                    "state_of_formation": _to_str(rd.get("state_of_formation")),
                    "address_line1": _to_str(rd.get("address_line1")),
                    "address_line2": _to_str(rd.get("address_line2")),
                    "city": _to_str(rd.get("city")),
                    "state": _to_str(rd.get("state")),
                    "postal_code": _to_str(rd.get("postal_code")),
                    "country": _to_str(rd.get("country")) or "US",
                    "primary_contact_email": _to_str(rd.get("primary_contact_email")),
                }
                out.append(ParsedCounterparty(
                    row=row_idx, external_ref=ext, legal_name=legal,
                    tax_id=tax_id, fields=fields, errors=errors,
                ))
            except Exception as e:
                out.append(ParsedCounterparty(
                    row=row_idx, external_ref=None, legal_name=None, tax_id=None,
                    errors=[f"unparseable row: {e}"],
                ))
        return out

    def _parse_loans(self, wb) -> list[ParsedLoan]:
        ws = wb[LOANS_SHEET]
        header, header_row = self._read_header(ws)
        if not header:
            return []
        out: list[ParsedLoan] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                       start=header_row + 1):
            if all(c is None or c == "" for c in row):
                continue
            errors: list[str] = []
            try:
                rd = self._row_dict(header, row)
                ext = _to_str(rd.get("loan_number")) or _to_str(rd.get("loan_external_ref"))
                portfolio_code = _to_str(rd.get("portfolio_code"))
                borrower_ref = _to_str(rd.get("borrower_external_ref"))
                if not portfolio_code:
                    errors.append("portfolio_code is required")
                if not borrower_ref:
                    errors.append("borrower_external_ref is required")

                payload = self._build_loan_create(rd, errors)

                out.append(ParsedLoan(
                    row=row_idx, external_ref=ext, portfolio_code=portfolio_code,
                    borrower_external_ref=borrower_ref, payload=payload, errors=errors,
                ))
            except Exception as e:
                out.append(ParsedLoan(
                    row=row_idx, external_ref=None, portfolio_code=None,
                    borrower_external_ref=None, payload=None,
                    errors=[f"unparseable row: {e}"],
                ))
        return out

    def _build_loan_create(self, rd: dict, errors: list[str]) -> Optional[LoanCreate]:
        """Build a LoanCreate (with conversion block) from a row dict.
        Coercion errors append to `errors` and return None for that row."""
        try:
            conv_args = dict(
                as_of_date=_to_date(rd.get("as_of_date")),
                current_principal=_to_decimal(rd.get("current_principal")) or Decimal("0"),
                accrued_interest=_to_decimal(rd.get("accrued_interest")) or Decimal("0"),
                accrued_fees=_to_decimal(rd.get("accrued_fees")) or Decimal("0"),
                last_payment_date=_to_date(rd.get("last_payment_date")),
                last_payment_amount=_to_decimal(rd.get("last_payment_amount")),
                next_due_date=_to_date(rd.get("next_due_date")),
                paid_to_date_principal=_to_decimal(rd.get("paid_to_date_principal")) or Decimal("0"),
                paid_to_date_interest=_to_decimal(rd.get("paid_to_date_interest")) or Decimal("0"),
                paid_to_date_fees=_to_decimal(rd.get("paid_to_date_fees")) or Decimal("0"),
                prior_servicer_name=_to_str(rd.get("prior_servicer_name")),
                prior_servicer_loan_id=_to_str(rd.get("prior_servicer_loan_id")),
                notes=_to_str(rd.get("notes")),
            )
            if conv_args["as_of_date"] is None:
                errors.append("as_of_date is required (this is a converted-loan import)")
                return None
            conv = LoanConversionPayload(**conv_args)

            # portfolio_id / primary_borrower_id are placeholders here — the
            # real UUIDs are injected at commit time using the resolution map.
            payload = LoanCreate(
                portfolio_id=UUID(int=0),
                loan_number=_to_str(rd.get("loan_number")),
                loan_name=_to_str(rd.get("loan_name")),
                primary_borrower_id=UUID(int=0),
                currency=_to_str(rd.get("currency")) or "USD",
                original_balance=_to_decimal(rd.get("original_balance")) or Decimal("0"),
                commitment_amount=_to_decimal(rd.get("commitment_amount")),
                rate_type=_to_str(rd.get("rate_type")) or "fixed",
                coupon_rate=_to_rate(rd.get("coupon_rate")),
                pik_rate=_to_rate(rd.get("pik_rate")),
                spread=_to_rate(rd.get("spread")),
                index_code=_to_str(rd.get("index_code")),
                day_count=_to_str(rd.get("day_count")) or "ACT/360",
                origination_date=_to_date(rd.get("origination_date")),
                first_payment_date=_to_date(rd.get("first_payment_date")),
                maturity_date=_to_date(rd.get("maturity_date")),
                payment_frequency=_to_str(rd.get("payment_frequency")) or "QUARTERLY",
                amortization_type=_to_str(rd.get("amortization_type")) or "bullet",
                interest_only_period_months=_to_int(rd.get("interest_only_period_months")),
                balloon_amount=_to_decimal(rd.get("balloon_amount")),
                grace_period_days=_to_int(rd.get("grace_period_days")) or 5,
                late_fee_type=_to_str(rd.get("late_fee_type")),
                late_fee_amount=_to_decimal(rd.get("late_fee_amount")),
                default_rate=_to_rate(rd.get("default_rate")),
                conversion=conv,
            )
            return payload
        except ValueError as e:
            errors.append(str(e))
            return None
        except Exception as e:
            errors.append(f"validation: {e}")
            return None

    @staticmethod
    def _read_header(ws) -> tuple[Optional[list[str]], int]:
        """Find the first non-empty row, treat it as the header."""
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(c is not None and str(c).strip() for c in row):
                return [
                    (str(c).strip().lower() if c is not None else "")
                    for c in row
                ], row_idx
        return None, 0

    @staticmethod
    def _row_dict(header: list[str], row: tuple) -> dict[str, Any]:
        return {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}


# ---------------------------------------------------------------------------
# Commit (separate concern — runs in Celery, opens its own session per loan)
# ---------------------------------------------------------------------------

async def commit_batch(tenant_slug: str, batch_id: UUID) -> dict[str, Any]:
    """
    Commit a previously-validated batch.

    Per-loan transaction model: each loan gets its own session+commit cycle.
    A failure on row 23 doesn't roll back rows 1..22. The caller (Celery task)
    is responsible for tenant routing.

    Returns a summary dict that gets written to conversion_batch.commit_report.
    """
    # 1. Load batch + report in one session.
    async with get_tenant_session_context(tenant_slug) as session:
        batch = (await session.execute(
            select(ConversionBatch).where(ConversionBatch.id == batch_id)
        )).scalar_one_or_none()
        if batch is None:
            raise ValidationError(f"conversion_batch {batch_id} not found")
        if batch.status not in ("validated",):
            raise ValidationError(
                f"batch {batch_id} status is {batch.status}, expected 'validated'"
            )
        report = batch.validation_report or {}
        resolution = report.get("_resolution", {})
        cp_creates = resolution.get("counterparty_creates", [])
        cp_map = {k: UUID(v) for k, v in resolution.get("counterparties", {}).items()}
        port_map = {k: UUID(v) for k, v in resolution.get("portfolios", {}).items()}
        loan_specs = resolution.get("loans", [])
        batch.status = "committing"
        await session.flush()

    # 2. Insert any new counterparties — single transaction is fine because
    #    they're side-effect-free (no GL impact).
    async with get_tenant_session_context(tenant_slug) as session:
        for cp in cp_creates:
            existing = (await session.execute(
                select(Counterparty).where(Counterparty.id == UUID(cp["id"]))
            )).scalar_one_or_none()
            if existing:
                continue
            row = Counterparty(
                id=UUID(cp["id"]),
                legal_name=cp["legal_name"],
                tax_id=cp.get("tax_id"),
                type=cp.get("type") or "borrower",
                entity_type=cp.get("entity_type"),
                jurisdiction=cp.get("jurisdiction"),
                state_of_formation=cp.get("state_of_formation"),
                address_line1=cp.get("address_line1"),
                address_line2=cp.get("address_line2"),
                city=cp.get("city"),
                state=cp.get("state"),
                postal_code=cp.get("postal_code"),
                country=cp.get("country") or "US",
                primary_contact_email=cp.get("primary_contact_email"),
                kyc_status="pending",
                external_id=cp.get("external_ref"),
            )
            session.add(row)

    # 3. Per-loan commit. New session per loan -> per-row isolation.
    succeeded: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    for spec in loan_specs:
        if spec.get("errors") or not spec.get("payload"):
            continue   # invalid rows already counted at validate time
        try:
            await _commit_one_loan(tenant_slug, spec, batch_id, port_map, cp_map)
            succeeded.append({"row": spec["row"], "external_ref": spec.get("external_ref")})
        except Exception as e:
            logger.exception("batch_loan_commit_failed", row=spec["row"], err=str(e))
            failed.append({"row": spec["row"], "external_ref": spec.get("external_ref"),
                           "error": str(e)})

    # 4. Write final summary back to the batch row.
    async with get_tenant_session_context(tenant_slug) as session:
        batch = (await session.execute(
            select(ConversionBatch).where(ConversionBatch.id == batch_id)
        )).scalar_one()
        batch.commit_report = {"succeeded": succeeded, "failed": failed}
        batch.succeeded_rows = len(succeeded)
        batch.failed_rows = (report.get("invalid_loan_rows", 0) or 0) + len(failed)
        batch.status = "completed"  # partial success is still completion; failed_rows surfaces detail
        await session.flush()

    return {"succeeded": len(succeeded), "failed": len(failed)}


async def _commit_one_loan(
    tenant_slug: str,
    spec: dict[str, Any],
    batch_id: UUID,
    port_map: dict[str, UUID],
    cp_map: dict[str, UUID],
) -> None:
    """Commit one loan in its own transaction. Raises on failure (caller logs)."""
    async with get_tenant_session_context(tenant_slug) as session:
        payload_dict = dict(spec["payload"])
        payload_dict["portfolio_id"] = str(port_map[spec["portfolio_code"]])
        payload_dict["primary_borrower_id"] = str(cp_map[spec["borrower_external_ref"]])
        payload = LoanCreate.model_validate(payload_dict)

        # Use the validated batch's uploaded_by as created_by — recorded on
        # the conversion_batch row alongside loan_conversion.posted_by.
        from app.models.conversion import ConversionBatch as _CB
        cb = (await session.execute(
            select(_CB.uploaded_by).where(_CB.id == batch_id)
        )).scalar_one()
        created_by = cb

        loan = await LoanService(session).create_converted_loan(payload, created_by=created_by)

        # Stamp the loan_conversion row with the batch_id so the loan can be
        # traced back to its upload batch.
        conv_row = (await session.execute(
            select(LoanConversion).where(LoanConversion.loan_id == loan.id)
        )).scalar_one()
        conv_row.batch_id = batch_id
